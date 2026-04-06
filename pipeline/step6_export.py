# ─────────────────────────────────────────────
#  step6_export.py
#  • Compile measurements from matched pairs
#  • Convert px → mm using ruler ratio
#  • Export to Excel with per-genotype sheets
# ─────────────────────────────────────────────

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings("ignore")


def compile_measurements(matches_left: list, matches_right: list,
                         px_per_mm_t1: float, px_per_mm_t2: float) -> pd.DataFrame:
    """
    Build a DataFrame with one row per matched pair.
    Columns:
      genotype_side, label, length_t1_px, length_t2_px,
      length_t1_mm, length_t2_mm, delta_mm, growth_rate_pct,
      matched, notes
    """
    rows = []

    for side_label, matches, in [("Left (A)", matches_left), ("Right (B)", matches_right)]:
        for m in matches:
            inst_t1 = m['t1']
            inst_t2 = m['t2']

            l_t1_px = inst_t1['length_px'] if inst_t1 else np.nan
            l_t2_px = inst_t2['length_px'] if inst_t2 else np.nan

            l_t1_mm = (l_t1_px / px_per_mm_t1) if (inst_t1 and px_per_mm_t1) else np.nan
            l_t2_mm = (l_t2_px / px_per_mm_t2) if (inst_t2 and px_per_mm_t2) else np.nan

            delta_mm = (l_t2_mm - l_t1_mm) if (not np.isnan(l_t1_mm) and not np.isnan(l_t2_mm)) else np.nan
            growth_pct = (delta_mm / l_t1_mm * 100) if (not np.isnan(delta_mm) and l_t1_mm > 0) else np.nan

            notes = []
            if not m['matched']:
                notes.append("unmatched")
            if inst_t1 and inst_t1.get('tangled'):
                notes.append("tangled-t1")
            if inst_t2 and inst_t2.get('tangled'):
                notes.append("tangled-t2")

            rows.append({
                'genotype_side'   : side_label,
                'label'           : m['label'],
                'length_t1_px'    : round(l_t1_px, 1) if not np.isnan(l_t1_px) else np.nan,
                'length_t2_px'    : round(l_t2_px, 1) if not np.isnan(l_t2_px) else np.nan,
                'length_t1_mm'    : round(l_t1_mm, 3) if not np.isnan(l_t1_mm) else np.nan,
                'length_t2_mm'    : round(l_t2_mm, 3) if not np.isnan(l_t2_mm) else np.nan,
                'delta_mm'        : round(delta_mm, 3) if not np.isnan(delta_mm) else np.nan,
                'growth_rate_pct' : round(growth_pct, 1) if not np.isnan(growth_pct) else np.nan,
                'matched'         : m['matched'],
                'notes'           : "; ".join(notes) if notes else "",
            })

    df = pd.DataFrame(rows)
    return df


def export_excel(df: pd.DataFrame, out_path: str,
                 px_per_mm_t1: float, px_per_mm_t2: float,
                 image_t1: str = "", image_t2: str = ""):
    """
    Write a nicely formatted Excel file with:
      - "All measurements" sheet
      - "Left (A)" sheet
      - "Right (B)" sheet
      - "Summary" sheet with means/SE
      - "Calibration" sheet with ruler info
    """
    out_path = Path(out_path)

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:

        # ── All measurements ──
        df.to_excel(writer, sheet_name="All measurements", index=False)
        _style_sheet(writer.sheets["All measurements"], df)

        # ── Per-side sheets ──
        for side in df['genotype_side'].unique():
            sub = df[df['genotype_side'] == side].reset_index(drop=True)
            sheet_name = side[:31]  # Excel limit
            sub.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name], sub)

        # ── Summary ──
        matched_df = df[df['matched']]
        summary_rows = []
        for side in df['genotype_side'].unique():
            s = matched_df[matched_df['genotype_side'] == side]
            if len(s) == 0:
                continue
            summary_rows.append({
                'genotype_side'         : side,
                'n'                     : len(s),
                'mean_length_t1_mm'     : s['length_t1_mm'].mean(),
                'se_length_t1_mm'       : s['length_t1_mm'].sem(),
                'mean_length_t2_mm'     : s['length_t2_mm'].mean(),
                'se_length_t2_mm'       : s['length_t2_mm'].sem(),
                'mean_delta_mm'         : s['delta_mm'].mean(),
                'se_delta_mm'           : s['delta_mm'].sem(),
                'mean_growth_rate_pct'  : s['growth_rate_pct'].mean(),
                'se_growth_rate_pct'    : s['growth_rate_pct'].sem(),
            })

        summary_df = pd.DataFrame(summary_rows).round(3)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        _style_sheet(writer.sheets["Summary"], summary_df)

        # ── Calibration ──
        cal_df = pd.DataFrame([
            {'parameter': 'Image t1', 'value': image_t1},
            {'parameter': 'Image t2', 'value': image_t2},
            {'parameter': 'px/mm t1', 'value': round(px_per_mm_t1, 3) if px_per_mm_t1 else 'N/A'},
            {'parameter': 'px/mm t2', 'value': round(px_per_mm_t2, 3) if px_per_mm_t2 else 'N/A'},
            {'parameter': 'Ruler tick spacing (mm)', 'value': 1.0},
        ])
        cal_df.to_excel(writer, sheet_name="Calibration", index=False)
        _style_sheet(writer.sheets["Calibration"], cal_df)

    print(f"  [export] Saved → {out_path}")
    return out_path


def _style_sheet(ws, df: pd.DataFrame):
    """Apply basic formatting to an openpyxl worksheet."""
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill    = PatternFill("solid", fgColor="DCE6F1")

    thin = Side(border_style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        # Auto-width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(str(col_name)) + 2)

    for row_idx in range(2, ws.max_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')


if __name__ == "__main__":
    # Quick test with fake data
    fake_matches_left = [
        {'label': 1, 'matched': True,
         't1': {'length_px': 250.0, 'tangled': False},
         't2': {'length_px': 310.0, 'tangled': False}},
        {'label': 2, 'matched': True,
         't1': {'length_px': 190.0, 'tangled': False},
         't2': {'length_px': 230.0, 'tangled': False}},
    ]
    fake_matches_right = [
        {'label': 3, 'matched': True,
         't1': {'length_px': 200.0, 'tangled': False},
         't2': {'length_px': 280.0, 'tangled': False}},
    ]
    df = compile_measurements(fake_matches_left, fake_matches_right, 45.0, 43.5)
    print(df)
    export_excel(df, "/tmp/test_output.xlsx", 45.0, 43.5,
                 image_t1="IMG_1714.JPG", image_t2="IMG_1728.JPG")
    print("Export OK")
